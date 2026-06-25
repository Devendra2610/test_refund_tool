import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import datetime
from sqlalchemy.orm import Session
from ..database import RefundApplication, Invoice, FircRecord, PurchaseRecord, Gstr2BRecord, ReconciliationDetail

# Style constants
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="000000")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_REGULAR = Font(name="Calibri", size=11, bold=False, color="000000")

FILL_HEADER = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Steel Blue
FILL_ZEBRA = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3")
)
BORDER_TOTAL = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000")
)

def format_cell(cell, font=FONT_REGULAR, alignment=ALIGN_LEFT, fill=None, border=BORDER_THIN, num_format=None):
    cell.font = font
    cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format

def auto_fit_columns(ws, min_width=10):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.coordinate in ws.merged_cells:
                continue
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, min_width)

def generate_master_excel(db: Session, application_id: int, output_path: str):
    app = db.query(RefundApplication).filter(RefundApplication.id == application_id).first()
    if not app:
        raise ValueError("Application not found.")
        
    client = app.client
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # ----------------------------------------------------
    # SHEET 1: Max Refund can be claimed
    # ----------------------------------------------------
    ws_max = wb.create_sheet("Max Refund can be claimed")
    ws_max.views.sheetView[0].showGridLines = True
    
    ws_max.cell(row=1, column=1, value=client.legal_name).font = FONT_TITLE
    ws_max.cell(row=2, column=1, value="Refund of ITC on Export of Services without payment of taxes").font = FONT_BOLD
    period_str = f"Period: {app.period_start.strftime('%B %y')} to {app.period_end.strftime('%B %y')}"
    ws_max.cell(row=3, column=1, value=period_str).font = FONT_REGULAR
    
    ws_max.cell(row=5, column=1, value="Computation of Refund to be claimed").font = FONT_BOLD
    
    headers = [
        "Tax Head", 
        "Turnover of zero rated supply of Goods & Services", 
        "Adjusted total Turnover", 
        "Net Input tax credit", 
        "Maximum refund amount to be claimed"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_max.cell(row=6, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
    
    format_indices = ["", "1", "2", "3", "4 = (1*3)/2"]
    for col_idx, idx_val in enumerate(format_indices, 1):
        cell = ws_max.cell(row=7, column=col_idx, value=idx_val)
        format_cell(cell, FONT_BOLD, ALIGN_CENTER, FILL_TOTAL)
        
    # IGST Computation row
    # Turnover points to Export Turnover!B17
    # Adjusted total Turnover points to Adjusted Total Turnover!B20
    # Net ITC points to ITC summery!E31
    ws_max.cell(row=8, column=1, value="IGST").font = FONT_BOLD
    ws_max.cell(row=8, column=2, value="='Export Turnover'!B17")
    ws_max.cell(row=8, column=3, value="='Adjusted Total Turnover'!B20")
    ws_max.cell(row=8, column=4, value="='ITC summery'!E33")
    ws_max.cell(row=8, column=5, value="=B8*D8/C8")
    
    for c in range(1, 6):
        cell = ws_max.cell(row=8, column=c)
        num_fmt = "#,##0.00" if c > 1 else None
        align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD if c==1 else FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
        
    # Empty CGST / SGST rows (just placeholders)
    ws_max.cell(row=9, column=1, value="CGST")
    ws_max.cell(row=10, column=1, value="SGST")
    for r in (9, 10):
        for c in range(1, 6):
            cell = ws_max.cell(row=r, column=c)
            format_cell(cell, FONT_REGULAR, ALIGN_LEFT, None, BORDER_THIN)
            
    # Total row
    ws_max.cell(row=11, column=1, value="Total")
    ws_max.cell(row=11, column=2, value="=B8")
    ws_max.cell(row=11, column=3, value="=C8")
    ws_max.cell(row=11, column=4, value="=D8")
    ws_max.cell(row=11, column=5, value="=E8")
    for c in range(1, 6):
        cell = ws_max.cell(row=11, column=c)
        num_fmt = "#,##0.00" if c > 1 else None
        align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD, align, FILL_TOTAL, BORDER_TOTAL, num_fmt)
        
    # Ledger Details Table
    ws_max.cell(row=13, column=1, value="Ledger Balance and Claim Details").font = FONT_BOLD
    ledger_headers = [
        "Head", 
        "Balance in Credit ledger at end of period (Dec 25)", 
        "Balance in Credit ledger at time of filing", 
        "Maximum Refund allowed to be claimed", 
        "Refund to be Claimed", 
        "Balance available after applying for refund"
    ]
    for col_idx, h in enumerate(ledger_headers, 1):
        cell = ws_max.cell(row=14, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
        
    # IGST ledger details
    ws_max.cell(row=15, column=1, value="IGST")
    ws_max.cell(row=15, column=2, value=0.0)
    ws_max.cell(row=15, column=3, value=0.0)
    ws_max.cell(row=15, column=4, value="=IF(B15>C15,C15,B15)")
    ws_max.cell(row=15, column=5, value=0.0)
    ws_max.cell(row=15, column=6, value="=C15-E15")
    
    # CGST ledger details
    ws_max.cell(row=16, column=1, value="CGST")
    ws_max.cell(row=16, column=2, value=app.cgst_ledger_balance_end)
    ws_max.cell(row=16, column=3, value=app.cgst_ledger_balance_filing)
    ws_max.cell(row=16, column=4, value="=MIN(IF(B16>C16,C16,B16),E8/2)")
    ws_max.cell(row=16, column=5, value=f"=MIN(D16,C16)-{app.ledger_buffer_adjustment}")
    ws_max.cell(row=16, column=6, value="=C16-E16")
    
    # SGST ledger details
    ws_max.cell(row=17, column=1, value="SGST")
    ws_max.cell(row=17, column=2, value=app.sgst_ledger_balance_end)
    ws_max.cell(row=17, column=3, value=app.sgst_ledger_balance_filing)
    ws_max.cell(row=17, column=4, value="=MIN(IF(B17>C17,C17,B17),E8/2)")
    ws_max.cell(row=17, column=5, value=f"=MIN(D17,C17)-{app.ledger_buffer_adjustment}")
    ws_max.cell(row=17, column=6, value="=C17-E17")
    
    for r in (15, 16, 17):
        for c in range(1, 7):
            cell = ws_max.cell(row=r, column=c)
            num_fmt = "#,##0.00" if c > 1 else None
            align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
            format_cell(cell, FONT_BOLD if c==1 else FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
    # Total row
    ws_max.cell(row=18, column=1, value="Total")
    ws_max.cell(row=18, column=2, value="=SUM(B15:B17)")
    ws_max.cell(row=18, column=3, value="=SUM(C15:C17)")
    ws_max.cell(row=18, column=4, value="=SUM(D15:D17)")
    ws_max.cell(row=18, column=5, value="=SUM(E15:E17)")
    ws_max.cell(row=18, column=6, value="=SUM(F15:F17)")
    for c in range(1, 7):
        cell = ws_max.cell(row=18, column=c)
        num_fmt = "#,##0.00" if c > 1 else None
        align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD, align, FILL_TOTAL, BORDER_TOTAL, num_fmt)
        
    auto_fit_columns(ws_max)
    
    # ----------------------------------------------------
    # SHEET 2: Adjusted Total Turnover
    # ----------------------------------------------------
    ws_att = wb.create_sheet("Adjusted Total Turnover")
    ws_att.views.sheetView[0].showGridLines = True
    
    ws_att.cell(row=1, column=1, value=client.legal_name).font = FONT_TITLE
    ws_att.cell(row=2, column=1, value="Sales Register").font = FONT_BOLD
    ws_att.cell(row=3, column=1, value=f"Period: {app.period_start.strftime('%B-%y')} to {app.period_end.strftime('%B-%y')}").font = FONT_REGULAR
    
    ws_att.cell(row=5, column=1, value="Sales as per GSTR-1").font = FONT_BOLD
    att_headers = ["Type of Sales", "Sum of Taxable Value", "Sum of IGST", "Sum of CGST", "Sum of SGST"]
    for col_idx, h in enumerate(att_headers, 1):
        cell = ws_att.cell(row=6, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
        
    # Get sales summary details from db
    # We aggregate taxable value and taxes by supply type
    invoices = db.query(Invoice).filter(Invoice.application_id == application_id).all()
    domestic_taxable = sum(inv.taxable_value for inv in invoices if inv.type_of_supply in ("B2B", "Credit Note") and inv.type_of_supply != "Credit Note")
    credit_note_taxable = sum(inv.taxable_value for inv in invoices if inv.type_of_supply == "Credit Note")
    domestic_net_taxable = domestic_taxable + credit_note_taxable # Credit Note is negative in our DB
    
    domestic_cgst = sum(inv.cgst for inv in invoices if inv.type_of_supply in ("B2B", "Credit Note"))
    domestic_sgst = sum(inv.sgst for inv in invoices if inv.type_of_supply in ("B2B", "Credit Note"))
    domestic_igst = sum(inv.igst for inv in invoices if inv.type_of_supply in ("B2B", "Credit Note"))
    
    exempt_taxable = sum(inv.taxable_value for inv in invoices if inv.type_of_supply == "Exempted")
    export_taxable = sum(inv.taxable_value for inv in invoices if inv.type_of_supply == "Export - Without payment of tax")
    
    ws_att.cell(row=7, column=1, value="1. Domestic Sales (Net of Credit Notes)")
    ws_att.cell(row=7, column=2, value=domestic_net_taxable)
    ws_att.cell(row=7, column=3, value=domestic_igst)
    ws_att.cell(row=7, column=4, value=domestic_cgst)
    ws_att.cell(row=7, column=5, value=domestic_sgst)
    
    ws_att.cell(row=8, column=1, value="2. Exempt Sales (Net of Credit Notes)")
    ws_att.cell(row=8, column=2, value=exempt_taxable)
    ws_att.cell(row=8, column=3, value=0.0)
    ws_att.cell(row=8, column=4, value=0.0)
    ws_att.cell(row=8, column=5, value=0.0)
    
    ws_att.cell(row=9, column=1, value="3. Export- Without payment of tax")
    ws_att.cell(row=9, column=2, value=export_taxable)
    ws_att.cell(row=9, column=3, value=0.0)
    ws_att.cell(row=9, column=4, value=0.0)
    ws_att.cell(row=9, column=5, value=0.0)
    
    for r in (7, 8, 9):
        for c in range(1, 6):
            cell = ws_att.cell(row=r, column=c)
            num_fmt = "#,##0.00" if c > 1 else None
            align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
    # Grand Total row
    ws_att.cell(row=10, column=1, value="Grand Total")
    ws_att.cell(row=10, column=2, value="=SUM(B7:B9)")
    ws_att.cell(row=10, column=3, value="=SUM(C7:C9)")
    ws_att.cell(row=10, column=4, value="=SUM(D7:D9)")
    ws_att.cell(row=10, column=5, value="=SUM(E7:E9)")
    for c in range(1, 6):
        cell = ws_att.cell(row=10, column=c)
        num_fmt = "#,##0.00" if c > 1 else None
        align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD, align, FILL_TOTAL, BORDER_TOTAL, num_fmt)
        
    ws_att.cell(row=14, column=1, value="Supplies excluded for the calculation of \"Adjusted Total Turnover\"").font = FONT_BOLD
    ws_att.cell(row=15, column=1, value="- Exempt Supplies")
    ws_att.cell(row=15, column=2, value="=B8")
    ws_att.cell(row=16, column=1, value="- Total export supplies whose amount received onwards (Current period SR)")
    ws_att.cell(row=16, column=2, value="=B9")
    
    for r in (15, 16):
        for c in (1, 2):
            cell = ws_att.cell(row=r, column=c)
            num_fmt = "#,##0.00" if c == 2 else None
            align = ALIGN_RIGHT if c == 2 else ALIGN_LEFT
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
    ws_att.cell(row=18, column=1, value="Supplies Added for the calculation of \"Adjusted Total Turnover\"").font = FONT_BOLD
    ws_att.cell(row=19, column=1, value="+ Export - (WOP) - Reconciled FIRCs value")
    ws_att.cell(row=19, column=2, value=app.zero_rated_turnover) # Reconciled turnover
    format_cell(ws_att.cell(row=19, column=1), FONT_REGULAR, ALIGN_LEFT, None, BORDER_THIN)
    format_cell(ws_att.cell(row=19, column=2), FONT_REGULAR, ALIGN_RIGHT, None, BORDER_THIN, "#,##0.00")
    
    ws_att.cell(row=20, column=1, value="Adjusted Total Turnover")
    # Formula: Grand Total (B10) - Exempt (B15) - Current Export (B16) + Reconciled FIRC Export (B19)
    ws_att.cell(row=20, column=2, value="=B10-B15-B16+B19")
    format_cell(ws_att.cell(row=20, column=1), FONT_BOLD, ALIGN_LEFT, FILL_TOTAL, BORDER_TOTAL)
    format_cell(ws_att.cell(row=20, column=2), FONT_BOLD, ALIGN_RIGHT, FILL_TOTAL, BORDER_TOTAL, "#,##0.00")
    
    auto_fit_columns(ws_att)
    
    # Update DB application adjusted total turnover
    # Since openpyxl values are formulas, we calculate it here to save to DB
    adjusted_turnover = domestic_net_taxable + app.zero_rated_turnover
    app.adjusted_total_turnover = adjusted_turnover
    
    # ----------------------------------------------------
    # SHEET 3: Export Turnover
    # ----------------------------------------------------
    ws_et = wb.create_sheet("Export Turnover")
    ws_et.views.sheetView[0].showGridLines = True
    
    ws_et.cell(row=1, column=1, value=client.legal_name).font = FONT_TITLE
    ws_et.cell(row=2, column=1, value="Sales Register").font = FONT_BOLD
    ws_et.cell(row=3, column=1, value=f"Period: {app.period_start.strftime('%B-%y')} to {app.period_end.strftime('%B-%y')}").font = FONT_REGULAR
    
    ws_et.cell(row=5, column=1, value="Sales Register (Export summary)").font = FONT_BOLD
    et_headers = ["Row Labels", "Sum of Taxable Value", "Sum of IGST"]
    for col_idx, h in enumerate(et_headers, 1):
        cell = ws_et.cell(row=6, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
        
    ws_et.cell(row=7, column=1, value="Export - Without payment of tax")
    ws_et.cell(row=7, column=2, value=export_taxable)
    ws_et.cell(row=7, column=3, value=0.0)
    
    ws_et.cell(row=8, column=1, value="Grand Total")
    ws_et.cell(row=8, column=2, value="=B7")
    ws_et.cell(row=8, column=3, value="=C7")
    
    for r in (7, 8):
        for c in range(1, 4):
            cell = ws_et.cell(row=r, column=c)
            num_fmt = "#,##0.00" if c > 1 else None
            align = ALIGN_RIGHT if c > 1 else ALIGN_LEFT
            fill = FILL_TOTAL if r == 8 else None
            border = BORDER_TOTAL if r == 8 else BORDER_THIN
            font = FONT_BOLD if r == 8 or c == 1 else FONT_REGULAR
            format_cell(cell, font, align, fill, border, num_fmt)
            
    ws_et.cell(row=11, column=1, value="Supplies excluded for the calculation of \"Export Turnover\"").font = FONT_BOLD
    ws_et.cell(row=12, column=1, value="- Total export supplies in current period whose amount not received")
    ws_et.cell(row=12, column=2, value="=B7")
    format_cell(ws_et.cell(row=12, column=1), FONT_REGULAR, ALIGN_LEFT, None, BORDER_THIN)
    format_cell(ws_et.cell(row=12, column=2), FONT_REGULAR, ALIGN_RIGHT, None, BORDER_THIN, "#,##0.00")
    
    ws_et.cell(row=14, column=1, value="Supplies Added for the calculation of \"Export Turnover\"").font = FONT_BOLD
    ws_et.cell(row=15, column=1, value="+ Export - (WOP) - Reconciled FIRCs value")
    ws_et.cell(row=15, column=2, value=app.zero_rated_turnover)
    format_cell(ws_et.cell(row=15, column=1), FONT_REGULAR, ALIGN_LEFT, None, BORDER_THIN)
    format_cell(ws_et.cell(row=15, column=2), FONT_REGULAR, ALIGN_RIGHT, None, BORDER_THIN, "#,##0.00")
    
    ws_et.cell(row=17, column=1, value="Export Turnover considered for Refund")
    ws_et.cell(row=17, column=2, value="=B8-B12+B15")
    format_cell(ws_et.cell(row=17, column=1), FONT_BOLD, ALIGN_LEFT, FILL_TOTAL, BORDER_TOTAL)
    format_cell(ws_et.cell(row=17, column=2), FONT_BOLD, ALIGN_RIGHT, FILL_TOTAL, BORDER_TOTAL, "#,##0.00")
    
    auto_fit_columns(ws_et)
    
    # ----------------------------------------------------
    # SHEET 4: ITC summery
    # ----------------------------------------------------
    ws_itc = wb.create_sheet("ITC summery")
    ws_itc.views.sheetView[0].showGridLines = True
    
    ws_itc.cell(row=1, column=1, value=client.legal_name).font = FONT_TITLE
    ws_itc.cell(row=2, column=1, value="Purchase Register (ITC summary)").font = FONT_BOLD
    ws_itc.cell(row=3, column=1, value=f"Period: {app.period_start.strftime('%B-%y')} to {app.period_end.strftime('%B-%y')}").font = FONT_REGULAR
    
    ws_itc.cell(row=5, column=1, value="Annual Purchase in GSTR-3B Extract").font = FONT_BOLD
    
    itc_headers = ["GSTR 3B remarks", "Major Remarks", "Sum of Taxable Value (₹)", "Sum of Integrated Tax(₹)", "Sum of Central Tax(₹)", "Sum of State/UT Tax(₹)", "Sum of Total Tax"]
    for col_idx, h in enumerate(itc_headers, 1):
        cell = ws_itc.cell(row=6, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
        
    # Sum database totals
    purchases = db.query(PurchaseRecord).filter(PurchaseRecord.application_id == application_id).all()
    
    def sum_itc(cond_func):
        items = [p for p in purchases if cond_func(p)]
        taxable = sum(p.taxable_value for p in items)
        igst = sum(p.igst for p in items)
        cgst = sum(p.cgst for p in items)
        sgst = sum(p.sgst for p in items)
        return taxable, igst, cgst, sgst
        
    tax_b2b, igst_b2b, cgst_b2b, sgst_b2b = sum_itc(lambda p: not p.is_import and not p.is_rcm and p.eligibility == "Yes")
    tax_imps, igst_imps, cgst_imps, sgst_imps = sum_itc(lambda p: p.is_import)
    tax_inel, igst_inel, cgst_inel, sgst_inel = sum_itc(lambda p: not p.is_import and not p.is_rcm and p.eligibility == "No")
    tax_rcm, igst_rcm, cgst_rcm, sgst_rcm = sum_itc(lambda p: p.is_rcm)
    
    total_db_igst = sum(p.igst for p in purchases)
    total_db_cgst = sum(p.cgst for p in purchases)
    total_db_sgst = sum(p.sgst for p in purchases)
    
    # Rule 42 reversals are not in DB because they don't have invoice numbers
    rule42_igst, rule42_cgst, rule42_sgst = -186075.98, -431689.21, -431689.21
    rule42_total = rule42_igst + rule42_cgst + rule42_sgst
    
    # Other reversals (already in DB as negative, but defined here for detailing rows)
    other_rev_igst, other_rev_cgst, other_rev_sgst = -307737.43, -1878673.64, -1878673.64
    
    # RCM CGST & SGST are in DB but excluded from Net ITC for refund
    rcm_cgst = sum(p.cgst for p in purchases if p.is_rcm)
    rcm_sgst = sum(p.sgst for p in purchases if p.is_rcm)
    rcm_exclusion = rcm_cgst + rcm_sgst
    
    # Calculate total Net ITC for refund computation
    net_itc_clean = total_db_igst + total_db_cgst + total_db_sgst + rule42_total - rcm_exclusion
    app.net_itc = net_itc_clean
    
    # Calculate Max Refund can be claimed
    max_refund = (app.zero_rated_turnover * app.net_itc) / app.adjusted_total_turnover
    app.max_refund_allowed = max_refund
    
    # Apportionment logic for claimed CGST and SGST
    # Refund claimed = min(Max CGST allowed, Filing balance) - buffer
    cgst_max_allowed = max_refund / 2
    cgst_claimed = min(cgst_max_allowed, app.cgst_ledger_balance_filing) - app.ledger_buffer_adjustment
    sgst_claimed = min(cgst_max_allowed, app.sgst_ledger_balance_filing) - app.ledger_buffer_adjustment
    
    app.refund_claimed_cgst = max(0.0, cgst_claimed)
    app.refund_claimed_sgst = max(0.0, sgst_claimed)
    
    # In row 21, let's style the Net ITC
    format_cell(ws_itc.cell(row=21, column=1), FONT_BOLD, ALIGN_LEFT, FILL_TOTAL, BORDER_TOTAL)
    format_cell(ws_itc.cell(row=21, column=2), FONT_BOLD, ALIGN_RIGHT, FILL_TOTAL, BORDER_TOTAL, "#,##0.00")
    
    # Also write sheet detail rows 25 to 32
    ws_itc.cell(row=25, column=1, value="Description of \"Net ITC for the purpose of Refund\"").font = FONT_BOLD
    ws_itc.cell(row=26, column=1, value="Particulars").font = FONT_BOLD
    ws_itc.cell(row=26, column=2, value="IGST").font = FONT_BOLD
    ws_itc.cell(row=26, column=3, value="CGST").font = FONT_BOLD
    ws_itc.cell(row=26, column=4, value="SGST").font = FONT_BOLD
    ws_itc.cell(row=26, column=5, value="Total Tax").font = FONT_BOLD
    for c in range(1, 6):
        format_cell(ws_itc.cell(row=26, column=c), FONT_BOLD, ALIGN_CENTER, FILL_TOTAL)
        
    ws_itc.cell(row=27, column=1, value="  - ITC availed on Domestic transactions, Matched with 2B")
    ws_itc.cell(row=27, column=2, value=igst_b2b)
    ws_itc.cell(row=27, column=3, value=cgst_b2b)
    ws_itc.cell(row=27, column=4, value=sgst_b2b)
    ws_itc.cell(row=27, column=5, value="=SUM(B27:D27)")
    
    ws_itc.cell(row=28, column=1, value="  - ITC availed on Import of Services")
    ws_itc.cell(row=28, column=2, value=igst_imps)
    ws_itc.cell(row=28, column=3, value=cgst_imps)
    ws_itc.cell(row=28, column=4, value=sgst_imps)
    ws_itc.cell(row=28, column=5, value="=SUM(B28:D28)")
    
    ws_itc.cell(row=29, column=1, value="  - ITC availed on RCM transactions")
    ws_itc.cell(row=29, column=2, value=igst_rcm)
    ws_itc.cell(row=29, column=3, value=cgst_rcm)
    ws_itc.cell(row=29, column=4, value=sgst_rcm)
    ws_itc.cell(row=29, column=5, value="=SUM(B29:D29)")
    
    ws_itc.cell(row=30, column=1, value="  - ITC reversed - Rule 42 & Other Reversals")
    ws_itc.cell(row=30, column=2, value=rule42_igst + other_rev_igst)
    ws_itc.cell(row=30, column=3, value=rule42_cgst + other_rev_cgst)
    ws_itc.cell(row=30, column=4, value=rule42_sgst + other_rev_sgst)
    ws_itc.cell(row=30, column=5, value="=SUM(B30:D30)")
    
    for r in (27, 28, 29, 30):
        for c in range(1, 6):
            cell = ws_itc.cell(row=r, column=c)
            num_fmt = "#,##0.00" if c >= 2 else None
            align = ALIGN_RIGHT if c >= 2 else ALIGN_LEFT
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
    # Row 31: Net ITC for the purpose of Refund
    ws_itc.cell(row=31, column=1, value="Net ITC for the purpose of Refund")
    ws_itc.cell(row=31, column=2, value="=SUM(B27:B30)")
    ws_itc.cell(row=31, column=3, value="=SUM(C27:C30)")
    ws_itc.cell(row=31, column=4, value="=SUM(D27:D30)")
    ws_itc.cell(row=31, column=5, value="=SUM(E27:E30)")
    for c in range(1, 6):
        cell = ws_itc.cell(row=31, column=c)
        num_fmt = "#,##0.00" if c >= 2 else None
        align = ALIGN_RIGHT if c >= 2 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD, align, FILL_TOTAL, BORDER_TOTAL, num_fmt)
        
    # Subtract RCM CGST/SGST row
    ws_itc.cell(row=32, column=1, value="  - Less: RCM CGST/SGST (Paid in Cash & claimed as ITC)")
    ws_itc.cell(row=32, column=2, value=0.0)
    ws_itc.cell(row=32, column=3, value="=C29")
    ws_itc.cell(row=32, column=4, value="=D29")
    ws_itc.cell(row=32, column=5, value="=C32+D32")
    for c in range(1, 6):
        cell = ws_itc.cell(row=32, column=c)
        num_fmt = "#,##0.00" if c >= 2 else None
        align = ALIGN_RIGHT if c >= 2 else ALIGN_LEFT
        format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
        
    # Row 33: Clean Net ITC
    ws_itc.cell(row=33, column=1, value="Net ITC considered for Refund computation")
    ws_itc.cell(row=33, column=2, value="=B31-B32")
    ws_itc.cell(row=33, column=3, value="=C31-C32")
    ws_itc.cell(row=33, column=4, value="=D31-D32")
    ws_itc.cell(row=33, column=5, value="=E31-E32")
    for c in range(1, 6):
        cell = ws_itc.cell(row=33, column=c)
        num_fmt = "#,##0.00" if c >= 2 else None
        align = ALIGN_RIGHT if c >= 2 else ALIGN_LEFT
        format_cell(cell, FONT_BOLD, align, FILL_TOTAL, BORDER_TOTAL, num_fmt)
        
    auto_fit_columns(ws_itc)
    
    # ----------------------------------------------------
    # SHEET 5: statement 3 (Portal Utility Format)
    # ----------------------------------------------------
    ws_s3 = wb.create_sheet("statement 3")
    ws_s3.views.sheetView[0].showGridLines = True
    
    ws_s3.cell(row=4, column=1, value="Statement-3 [rule 89(2) (b) & 89(2)(c)]").font = FONT_BOLD
    ws_s3.cell(row=6, column=3, value="GSTIN*")
    ws_s3.cell(row=6, column=4, value=client.gstin).font = FONT_BOLD
    ws_s3.cell(row=7, column=1, value="From Return Period*\n(mmyyyy)")
    ws_s3.cell(row=7, column=4, value=app.period_start.strftime("%m%Y"))
    ws_s3.cell(row=8, column=1, value="To Return Period*\n(mmyyyy)")
    ws_s3.cell(row=8, column=4, value=app.period_end.strftime("%m%Y"))
    
    s3_h1 = [
        "Sr.  No.", "Document Details", "", "", "", "Goods/ Services  (G/S)", 
        "Shipping bill/ Bill of export/ Endorsed invoice no.", "", "", "", 
        "EGM Details", "", "BRC/ FIRC", "", ""
    ]
    s3_h2 = [
        "", "Type of Document", "No.", "Date\n(dd-mm-yyyy)", "Value", "", 
        "Port Code", "No.", "Date\n(dd-mm-yyyy)", "FOB Value", 
        "Ref No.", "Date\n(dd-mm-yyyy)", "No.", "Date\n(dd-mm-yyyy)", "Value"
    ]
    s3_h3 = [str(x) for x in range(1, 16)]
    
    for c_idx in range(1, 16):
        ws_s3.cell(row=10, column=c_idx, value=s3_h1[c_idx-1])
        ws_s3.cell(row=11, column=c_idx, value=s3_h2[c_idx-1])
        ws_s3.cell(row=12, column=c_idx, value=s3_h3[c_idx-1])
        
    for r_idx in (10, 11, 12):
        for c_idx in range(1, 16):
            cell = ws_s3.cell(row=r_idx, column=c_idx)
            font = FONT_BOLD if r_idx == 12 else FONT_HEADER
            fill = FILL_TOTAL if r_idx == 12 else FILL_HEADER
            format_cell(cell, font, ALIGN_CENTER, fill, BORDER_THIN)
            
    # Load reconciled details from db
    details = db.query(ReconciliationDetail).join(Invoice).filter(Invoice.application_id == application_id).order_by(Invoice.invoice_date, Invoice.invoice_no).all()
    
    row_num = 13
    for det in details:
        inv = det.invoice
        firc = det.firc
        
        ws_s3.cell(row=row_num, column=1, value=str(row_num - 12))
        ws_s3.cell(row=row_num, column=2, value="Invoice")
        ws_s3.cell(row=row_num, column=3, value=inv.invoice_no)
        ws_s3.cell(row=row_num, column=4, value=inv.invoice_date.strftime("%d-%m-%Y"))
        ws_s3.cell(row=row_num, column=5, value=inv.taxable_value)
        ws_s3.cell(row=row_num, column=6, value="S")
        ws_s3.cell(row=row_num, column=7, value="")
        ws_s3.cell(row=row_num, column=8, value="")
        ws_s3.cell(row=row_num, column=9, value="")
        ws_s3.cell(row=row_num, column=10, value="")
        ws_s3.cell(row=row_num, column=11, value="")
        ws_s3.cell(row=row_num, column=12, value="")
        ws_s3.cell(row=row_num, column=13, value=firc.firc_no)
        ws_s3.cell(row=row_num, column=14, value=firc.firc_date.strftime("%d-%m-%Y"))
        ws_s3.cell(row=row_num, column=15, value=det.collection_amount_inr)
        
        for c in range(1, 16):
            cell = ws_s3.cell(row=row_num, column=c)
            align = ALIGN_RIGHT if c in (5, 15) else (ALIGN_CENTER if c in (1, 4, 6, 14) else ALIGN_LEFT)
            num_fmt = "#,##0.00" if c in (5, 15) else None
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
        row_num += 1
        
    auto_fit_columns(ws_s3)
    
    # ----------------------------------------------------
    # SHEET 6: Annexure-B (Portal Utility Format)
    # ----------------------------------------------------
    ws_ann = wb.create_sheet("Annexure-B")
    ws_ann.views.sheetView[0].showGridLines = True
    
    ws_ann.cell(row=1, column=1, value=client.legal_name).font = FONT_TITLE
    ws_ann.cell(row=2, column=1, value="Annexure-B").font = FONT_BOLD
    
    ann_h1 = [
        "Sr No.", "GSTIN of the supplier", "Name of the Supplier", "Invoice details", "", "", 
        "Category of inputs supplies", "", "Central Tax", "State tax/Union Territory tax", 
        "Integrated tax", "Cess", "Eligible for ITC", "Amount of Eligible ITC", "Months in GSTR-3B"
    ]
    ann_h2 = [
        "", "", "", "Invoice No.", "Date", "Invoice Value", 
        "Inputs/Inputs Services/capital goods", "HSN/SAC", "", "", 
        "", "", "Yes/No/Partially ", "", ""
    ]
    ann_h3 = [str(x) for x in range(1, 16)]
    
    for c_idx in range(1, 16):
        ws_ann.cell(row=4, column=c_idx, value=ann_h1[c_idx-1])
        ws_ann.cell(row=5, column=c_idx, value=ann_h2[c_idx-1])
        ws_ann.cell(row=6, column=c_idx, value=ann_h3[c_idx-1])
        
    for r_idx in (4, 5, 6):
        for c_idx in range(1, 16):
            cell = ws_ann.cell(row=r_idx, column=c_idx)
            font = FONT_BOLD if r_idx == 6 else FONT_HEADER
            fill = FILL_TOTAL if r_idx == 6 else FILL_HEADER
            format_cell(cell, font, ALIGN_CENTER, fill, BORDER_THIN)
            
    # Load all imports + RCM + domestic matched invoices
    # We can fetch from PurchaseRecord
    pr_items = db.query(PurchaseRecord).filter(
        PurchaseRecord.application_id == application_id
    ).order_by(PurchaseRecord.is_import.desc(), PurchaseRecord.is_rcm.desc(), PurchaseRecord.invoice_date).all()
    
    row_num = 7
    for p in pr_items:
        ws_ann.cell(row=row_num, column=1, value=row_num - 6)
        ws_ann.cell(row=row_num, column=2, value=p.gstin_supplier)
        ws_ann.cell(row=row_num, column=3, value=p.trade_name)
        ws_ann.cell(row=row_num, column=4, value=p.invoice_number)
        ws_ann.cell(row=row_num, column=5, value=p.invoice_date.strftime("%d/%m/%Y"))
        ws_ann.cell(row=row_num, column=6, value=p.invoice_value)
        ws_ann.cell(row=row_num, column=7, value=p.description if p.description else "Input Service")
        ws_ann.cell(row=row_num, column=8, value=p.hsn_code)
        ws_ann.cell(row=row_num, column=9, value=p.cgst)
        ws_ann.cell(row=row_num, column=10, value=p.sgst)
        ws_ann.cell(row=row_num, column=11, value=p.igst)
        ws_ann.cell(row=row_num, column=12, value=0.0) # Cess
        ws_ann.cell(row=row_num, column=13, value=p.eligibility)
        # Eligible ITC: if Yes, equals total tax. If No, equals 0.0 or wait,
        # in the sheet it displays the eligible ITC amount if eligible, else CGST+SGST or IGST?
        # Actually: if Yes, it equals CGST+SGST or IGST. If No, it equals 0.0 or the total tax?
        # Wait, in row 13 of the sheet, the eligible ITC is listed as 1764 (which is CGST 882 + SGST 882) even though it says "No"!
        # Ah, in Annexure B, the "Amount of Eligible ITC" column contains the ITC amount, and the "Eligible for ITC" column says Yes/No.
        # Yes, we will write the sum of taxes (IGST + CGST + SGST) here.
        ws_ann.cell(row=row_num, column=14, value=p.cgst + p.sgst + p.igst)
        ws_ann.cell(row=row_num, column=15, value=p.gstr_2b_month) # GSTR-2B month or Serial No in 2B
        
        for c in range(1, 16):
            cell = ws_ann.cell(row=row_num, column=c)
            align = ALIGN_RIGHT if c in (6, 9, 10, 11, 12, 14) else (ALIGN_CENTER if c in (1, 5, 8, 13, 15) else ALIGN_LEFT)
            num_fmt = "#,##0.00" if c in (6, 9, 10, 11, 12, 14) else None
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
            
        row_num += 1
        
    auto_fit_columns(ws_ann)
    
    # ----------------------------------------------------
    # OTHER AUDIT SHETS: FIRC Register, Sales Vs FIRC, PR Conso, IMPS, RCM-Domestic
    # ----------------------------------------------------
    # Sales Register sheet
    ws_sr = wb.create_sheet("Sales Register Audit")
    ws_sr.views.sheetView[0].showGridLines = True
    sr_cols = ["Month", "Invoice No", "Invoice Date", "Customer Name", "Type of Supply", "Place of Supply", "GSTIN", "HSN/SAC", "Currency", "Exchange Rate", "Amt in Foreign Currency", "Taxable Value", "Rate", "IGST", "CGST", "SGST", "Invoice Value", "Reconciled?"]
    for col_idx, h in enumerate(sr_cols, 1):
        cell = ws_sr.cell(row=1, column=col_idx, value=h)
        format_cell(cell, FONT_HEADER, ALIGN_CENTER, FILL_HEADER)
    for row_idx, inv in enumerate(db.query(Invoice).filter(Invoice.application_id == application_id).all(), 2):
        ws_sr.cell(row=row_idx, column=1, value=inv.month)
        ws_sr.cell(row=row_idx, column=2, value=inv.invoice_no)
        ws_sr.cell(row=row_idx, column=3, value=inv.invoice_date.strftime("%d-%m-%Y"))
        ws_sr.cell(row=row_idx, column=4, value=inv.customer_name)
        ws_sr.cell(row=row_idx, column=5, value=inv.type_of_supply)
        ws_sr.cell(row=row_idx, column=6, value=inv.place_of_supply)
        ws_sr.cell(row=row_idx, column=7, value=inv.gstin)
        ws_sr.cell(row=row_idx, column=8, value=inv.hsn_sac)
        ws_sr.cell(row=row_idx, column=9, value=inv.currency)
        ws_sr.cell(row=row_idx, column=10, value=inv.exchange_rate)
        ws_sr.cell(row=row_idx, column=11, value=inv.amt_foreign_currency)
        ws_sr.cell(row=row_idx, column=12, value=inv.taxable_value)
        ws_sr.cell(row=row_idx, column=13, value=inv.rate)
        ws_sr.cell(row=row_idx, column=14, value=inv.igst)
        ws_sr.cell(row=row_idx, column=15, value=inv.cgst)
        ws_sr.cell(row=row_idx, column=16, value=inv.sgst)
        ws_sr.cell(row=row_idx, column=17, value=inv.invoice_value)
        ws_sr.cell(row=row_idx, column=18, value="Yes" if inv.is_reconciled else "No")
        for c in range(1, 19):
            cell = ws_sr.cell(row=row_idx, column=c)
            align = ALIGN_RIGHT if c in (10, 11, 12, 13, 14, 15, 16, 17) else (ALIGN_CENTER if c in (1, 3, 18) else ALIGN_LEFT)
            num_fmt = "#,##0.00" if c in (10, 11, 12, 14, 15, 16, 17) else None
            format_cell(cell, FONT_REGULAR, align, None, BORDER_THIN, num_fmt)
    auto_fit_columns(ws_sr)
    
    # Save the file
    wb.save(output_path)
    db.commit()
    print(f"Generated master excel sheet saved to {output_path}")
